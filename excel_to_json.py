import os
import json
import openpyxl
import warnings

# Suppress warnings from openpyxl about default styles
warnings.filterwarnings("ignore")

# Configuration
SOURCE_DIR = r"c:\Users\jalpa\Desktop\atlas\ep-ppt\wetransfer_all-media-data_2025-11-20_0937\report data for all media"
OUTPUT_FILE = r"c:\Users\jalpa\Desktop\atlas\ep-ppt\output_data.json"
REQUIRED_COLS = ["SUPER CATEGORY", "PRODUCT GROUP"]

def normalize_header(header):
    if header is None:
        return ""
    return str(header).strip().upper()

def process_file(file_path):
    """
    Process a single Excel file and return a dict of {SuperCategory: Set(ProductGroups)}.
    Using openpyxl read_only=True to handle malformed dimensions.
    """
    local_data = {}
    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        # We only look at the first/active sheet usually
        try:
             ws = wb.active
        except:
             # If active sheet fails, try first sheet
             if wb.sheetnames:
                 ws = wb[wb.sheetnames[0]]
             else:
                 return local_data

        rows = ws.iter_rows(values_only=True)
        
        # scan for headers in first 20 rows
        header_map = {} # Col Name -> Index
        found_headers = False
        
        # We need to buffer the row iterator because once consumed it's gone
        # But for read_only, we can just iterate.
        
        row_idx = 0
        for row in rows:
            row_idx += 1
            if row_idx > 50: # If headers not found in first 50 rows, probably wrong sheet
                break
                
            # Check row for headers
            current_row_headers = [normalize_header(c) for c in row]
            
            # Check if this row contains our required columns
            if all(req in current_row_headers for req in REQUIRED_COLS):
                # Build map
                for idx, col_name in enumerate(current_row_headers):
                    if col_name in REQUIRED_COLS:
                        header_map[col_name] = idx
                found_headers = True
                break
        
        if found_headers:
            # Continue iterating the REST of the rows
            sc_idx = header_map["SUPER CATEGORY"]
            pg_idx = header_map["PRODUCT GROUP"]
            
            for row in rows:
                # bounds check
                if len(row) <= max(sc_idx, pg_idx):
                    continue
                    
                sc_val = row[sc_idx]
                pg_val = row[pg_idx]
                
                if sc_val and pg_val:
                    sc = str(sc_val).strip()
                    pg = str(pg_val).strip()
                    
                    if sc and pg and sc.lower() != "super category": # Skip if header repeated
                        if sc not in local_data:
                            local_data[sc] = set()
                        local_data[sc].add(pg)
        else:
            print(f"  [WARN] Headers {REQUIRED_COLS} not found in first 50 rows.")

    except Exception as e:
        print(f"  [ERROR] Processing failed: {e}")
    finally:
        if wb:
            try:
                wb.close()
            except:
                pass
                
    return local_data

def extract_data():
    aggregated_data = {} # Key: Super Category, Value: Set of Product Groups
    
    print(f"Scanning directory: {SOURCE_DIR}")
    
    files_processed = 0
    files_skipped = 0
    
    for root, dirs, files in os.walk(SOURCE_DIR):
        for file in files:
            if file.lower().endswith(".xlsx") and not file.startswith("~$"):
                file_path = os.path.join(root, file)
                print(f"Processing: {file}")
                
                data = process_file(file_path)
                if data:
                    files_processed += 1
                    # Merge data
                    for sc, pgs in data.items():
                        if sc not in aggregated_data:
                            aggregated_data[sc] = set()
                        aggregated_data[sc].update(pgs)
                else:
                    files_skipped += 1

    # Convert to list of dictionaries for JSON output
    output_list = []
    for super_cat, prod_groups in aggregated_data.items():
        output_list.append({
            "super_category": super_cat,
            "product_groups_array": sorted(list(prod_groups))
        })
    
    # Sort output for consistent results
    output_list.sort(key=lambda x: x["super_category"])

    # Write to JSON
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(output_list, f, indent=4, ensure_ascii=False)
    
    print("-" * 30)
    print(f"Processing Complete.")
    print(f"Files Processed (with data): {files_processed}")
    print(f"Unique Super Categories Found: {len(output_list)}")
    print(f"Output saved to: {OUTPUT_FILE}")

if __name__ == "__main__":
    extract_data()
