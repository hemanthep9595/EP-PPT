
import openpyxl
import pandas as pd
import sys

FILE_PATH = r"c:\Users\jalpa\Desktop\atlas\ep-ppt\wetransfer_all-media-data_2025-11-20_0937\report data for all media\Digital\Aug -25.xlsx"

print(f"Testing file: {FILE_PATH}")

print("\n--- Test 1: openpyxl regular load ---")
try:
    wb = openpyxl.load_workbook(FILE_PATH)
    print("Success!")
except Exception as e:
    print(f"Failed: {e}")

print("\n--- Test 2: openpyxl read_only=True ---")
try:
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True)
    print("Success!")
    # Try to access a sheet
    ws = wb.active
    print(f"Active sheet: {ws.title}")
    # Read first few rows
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        print(f"Row {i+1}: {row}")
except Exception as e:
    print(f"Failed: {e}")

print("\n--- Test 3: pandas with engine='openpyxl' ---")
try:
    df = pd.read_excel(FILE_PATH, engine='openpyxl')
    print("Success!")
except Exception as e:
    print(f"Failed: {e}")
